"""Attachment text extraction.

Port of the TypeScript extractor's document handling: pdf-parse → pypdf,
mammoth → python-docx, xlsx → openpyxl. Behaviour is deliberately the same,
including the silent-skip contract: any unsupported format or parse failure
returns None and the attachment is simply left out of the prompt. A broken
attachment must never fail an extraction.
"""

from __future__ import annotations

import csv
import datetime as dt
import io
from typing import Any

XLSX_MIME_TYPES = frozenset(
    {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "application/x-excel",
    }
)

DOCX_MIME_TYPES = frozenset(
    {
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    }
)

PLAIN_MIME_TYPES = frozenset({"text/csv", "text/plain"})


def _cell_to_text(value: Any) -> str:
    """Render one spreadsheet cell the way SheetJS's sheet_to_csv would."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, dt.datetime):
        return value.date().isoformat() if value.time() == dt.time.min else value.isoformat(sep=" ")
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _sheet_to_csv(rows: list[tuple[Any, ...]]) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    for row in rows:
        writer.writerow([_cell_to_text(c) for c in row])
    return buf.getvalue()


def _pdf_text(data: bytes, original_name: str) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    text = "\n".join(page.extract_text() or "" for page in reader.pages).strip()
    if text:
        return text
    # No text layer — likely a scanned image PDF
    return (
        f"[Scanned PDF — no text layer detected. "
        f'A reviewer should inspect "{original_name}" directly.]'
    )


def _xlsx_text(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        blocks = []
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            blocks.append(f"[Sheet: {ws.title}]\n{_sheet_to_csv(rows)}")
        return "\n\n".join(blocks)
    finally:
        wb.close()


def _docx_text(data: bytes) -> str | None:
    from docx import Document

    doc = Document(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs).strip() or None


def extract_document_text(
    data: bytes,
    mime_type: str,
    original_name: str,
) -> str | None:
    """Extract plain text from an attachment, or None if it cannot be read."""
    try:
        if mime_type == "application/pdf":
            return _pdf_text(data, original_name)

        if mime_type in XLSX_MIME_TYPES:
            return _xlsx_text(data)

        if mime_type in DOCX_MIME_TYPES:
            return _docx_text(data)

        if mime_type in PLAIN_MIME_TYPES:
            return data.decode("utf-8", errors="replace").strip() or None

        return None  # unsupported format — silently skip
    except Exception:
        return None
