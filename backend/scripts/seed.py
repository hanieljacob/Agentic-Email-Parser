"""Reset the database to the demo state.

Two steps, so that one command gives a reviewer something to look at:

  1. Reload the canonical tables from backend/data/db.xlsx, mapping the
     workbook's integer ids onto UUIDs.
  2. Ingest the committed fixture emails through the real pipeline, which
     populates the review queue.

This is destructive by design — it truncates the pipeline tables too, so
running it twice produces exactly the same state rather than stacking a
second set of proposed changes on top of the first. It is a development and
demo command; nothing calls it at runtime.

    python -m backend.scripts.seed              # canonical data + fixtures
    python -m backend.scripts.seed --no-fixtures
"""

from __future__ import annotations

import argparse
import asyncio
import datetime as dt
import os
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from backend.config import PROJECT_ROOT
from backend.db import close_pool, connection, open_pool


def _rows(worksheet: Any) -> list[dict[str, Any]]:
    """Read a sheet as dicts keyed by its header row, skipping blank rows."""
    it = worksheet.iter_rows(values_only=True)
    try:
        header = [str(h) if h is not None else "" for h in next(it)]
    except StopIteration:
        return []

    out: list[dict[str, Any]] = []
    for row in it:
        if all(cell is None for cell in row):
            continue
        out.append({key: value for key, value in zip(header, row) if key})
    return out


def _int(value: Any) -> int | None:
    """Workbook integers arrive as floats (1.0); normalise them."""
    return None if value is None else int(value)


def _date(value: Any) -> dt.date | None:
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return None


def _nullify(value: Any) -> Any:
    """The workbook uses '' and '-' to mean "no value"."""
    if value is None or value == "" or value == "-":
        return None
    return value


async def seed() -> None:
    xlsx_path = Path(
        os.environ.get("XLSX_PATH", PROJECT_ROOT / "backend" / "data" / "db.xlsx")
    )
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)

    products = _rows(workbook["product"])
    suppliers = _rows(workbook["supplier"])
    purchase_orders = _rows(workbook["purchase_order"])
    lines = _rows(workbook["purchase_order_line"])
    supplier_products = _rows(workbook["supplier_product"])
    workbook.close()

    print(
        f"xlsx loaded: {len(products)} products, {len(suppliers)} suppliers, "
        f"{len(purchase_orders)} POs, {len(lines)} lines, "
        f"{len(supplier_products)} supplier_products"
    )

    async with connection() as conn:
        async with conn.transaction():
            # Truncate in reverse FK dependency order; CASCADE handles children.
            # The pipeline tables go too: proposed_changes reference canonical
            # rows by a logical FK that Postgres cannot cascade, so leaving
            # them behind would strand them against ids that no longer exist —
            # and re-running the fixtures would stack a duplicate queue.
            await conn.execute(
                """
                TRUNCATE audit_log, proposed_changes, extraction_runs,
                         email_attachments, emails, supplier_corrections,
                         supplier_email_aliases, supplier_product,
                         purchase_order_line, purchase_order,
                         product, supplier
                RESTART IDENTITY CASCADE
                """
            )

            product_ids: dict[int, str] = {}
            supplier_ids: dict[int, str] = {}
            po_ids: dict[int, str] = {}

            for row in products:
                cur = await conn.execute(
                    "INSERT INTO product (legacy_id, sku, title) VALUES (%s,%s,%s) RETURNING id",
                    (_int(row["id"]), row["sku"], _nullify(row["title"])),
                )
                product_ids[_int(row["id"])] = (await cur.fetchone())["id"]
            print(f"  ✓ {len(products)} products")

            for row in suppliers:
                cur = await conn.execute(
                    "INSERT INTO supplier (legacy_id, name, email) VALUES (%s,%s,%s) RETURNING id",
                    (_int(row["id"]), row["name"], row["email"]),
                )
                supplier_id = (await cur.fetchone())["id"]
                supplier_ids[_int(row["id"])] = supplier_id

                # Seed the primary email as an alias so alias resolution works too.
                await conn.execute(
                    "INSERT INTO supplier_email_aliases (supplier_id, email_address) VALUES (%s,%s)",
                    (supplier_id, row["email"]),
                )
            print(f"  ✓ {len(suppliers)} suppliers (+ {len(suppliers)} email aliases)")

            for row in purchase_orders:
                supplier_id = supplier_ids.get(_int(row["supplier_id"]))
                if supplier_id is None:
                    raise ValueError(f"Unknown supplier legacy_id: {row['supplier_id']}")
                cur = await conn.execute(
                    """
                    INSERT INTO purchase_order
                      (legacy_id, reference_num, supplier_id, delivery_date)
                    VALUES (%s,%s,%s,%s) RETURNING id
                    """,
                    (
                        _int(row["id"]),
                        row["reference_num"],
                        supplier_id,
                        _date(row["delivery_date"]),
                    ),
                )
                po_ids[_int(row["id"])] = (await cur.fetchone())["id"]
            print(f"  ✓ {len(purchase_orders)} purchase_orders")

            for row in lines:
                po_id = po_ids.get(_int(row["purchase_order_id"]))
                product_id = product_ids.get(_int(row["product_id"]))
                if po_id is None:
                    raise ValueError(f"Unknown PO legacy_id: {row['purchase_order_id']}")
                if product_id is None:
                    raise ValueError(f"Unknown product legacy_id: {row['product_id']}")
                await conn.execute(
                    """
                    INSERT INTO purchase_order_line
                      (legacy_id, purchase_order_id, product_id, quantity, delivery_date)
                    VALUES (%s,%s,%s,%s,%s)
                    """,
                    (
                        _int(row["id"]),
                        po_id,
                        product_id,
                        row["quantity"],
                        _date(row["delivery_date"]),
                    ),
                )
            print(f"  ✓ {len(lines)} purchase_order_lines")

            for row in supplier_products:
                supplier_id = supplier_ids.get(_int(row["supplier_id"]))
                product_id = product_ids.get(_int(row["product_id"]))
                if supplier_id is None or product_id is None:
                    raise ValueError(f"Unknown ids in supplier_product row: {row}")
                await conn.execute(
                    """
                    INSERT INTO supplier_product
                      (supplier_id, product_id, supplier_sku, price_per_unit)
                    VALUES (%s,%s,%s,%s)
                    """,
                    (
                        supplier_id,
                        product_id,
                        _nullify(row["supplier_sku"]),
                        _nullify(row["price_per_unit"]),
                    ),
                )
            print(f"  ✓ {len(supplier_products)} supplier_products")

    print("Canonical tables reloaded.")


async def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--no-fixtures",
        action="store_true",
        help="load canonical data only, leaving the review queue empty",
    )
    args = parser.parse_args()

    await open_pool()
    try:
        await seed()
        if not args.no_fixtures:
            from backend.scripts.load_fixtures import load_fixtures

            print("Ingesting fixture emails through the pipeline...")
            totals = await load_fixtures()
            print(
                f"  {totals['proposed']} changes proposed, "
                f"{totals['auto_applied']} auto-applied, "
                f"{totals['pending']} awaiting review."
            )
    finally:
        await close_pool()


if __name__ == "__main__":
    asyncio.run(main())
